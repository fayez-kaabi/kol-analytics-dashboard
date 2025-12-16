"""
Excel file parser for KOL data.

BONUS FEATURE: Excel parsing instead of JSON loading.

This module handles reading .xlsx files and converting them to the KOL data format.
Supports automatic column mapping and data type conversion.
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Parser for Excel files containing KOL data.
    
    Handles:
    - Column name mapping (Excel headers -> JSON field names)
    - Data type conversion
    - Missing value handling
    - Multiple sheet support
    """
    
    # Mapping of possible Excel column names to our JSON field names
    COLUMN_MAPPINGS = {
        # ID fields
        'id': 'id',
        'kol_id': 'id',
        'identifier': 'id',
        
        # Name fields
        'name': 'name',
        'full_name': 'name',
        'kol_name': 'name',
        'doctor_name': 'name',
        
        # Affiliation fields
        'affiliation': 'affiliation',
        'institution': 'affiliation',
        'organization': 'affiliation',
        'hospital': 'affiliation',
        'university': 'affiliation',
        
        # Country fields
        'country': 'country',
        'nation': 'country',
        
        # City fields
        'city': 'city',
        'location': 'city',
        
        # Expertise fields
        'expertise_area': 'expertiseArea',
        'expertisearea': 'expertiseArea',
        'expertise': 'expertiseArea',
        'specialization': 'expertiseArea',
        'specialty': 'expertiseArea',
        'field': 'expertiseArea',
        
        # Publications fields
        'publications_count': 'publicationsCount',
        'publicationscount': 'publicationsCount',
        'publications': 'publicationsCount',
        'num_publications': 'publicationsCount',
        'publication_count': 'publicationsCount',
        
        # H-Index fields
        'h_index': 'hIndex',
        'hindex': 'hIndex',
        'h-index': 'hIndex',
        
        # Citations fields
        'citations': 'citations',
        'total_citations': 'citations',
        'citation_count': 'citations',
    }
    
    @classmethod
    def parse_excel_file(cls, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse an Excel file and return list of KOL records.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            List of dictionaries with KOL data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is invalid
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        logger.info(f"Parsing Excel file: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Try to find the data sheet (use first sheet or look for specific names)
            sheet = cls._find_data_sheet(workbook)
            
            # Parse the sheet
            records = cls._parse_sheet(sheet)
            
            logger.info(f"Successfully parsed {len(records)} records from Excel")
            return records
            
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            raise ValueError(f"Invalid Excel file format: {e}")
    
    @classmethod
    def _find_data_sheet(cls, workbook: openpyxl.Workbook) -> Worksheet:
        """Find the worksheet containing KOL data."""
        # Look for sheets with common names
        preferred_names = ['kol', 'data', 'sheet1', 'kols', 'doctors']
        
        for name in preferred_names:
            for sheet_name in workbook.sheetnames:
                if name in sheet_name.lower():
                    return workbook[sheet_name]
        
        # Fallback to first sheet
        return workbook.active
    
    @classmethod
    def _parse_sheet(cls, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse a worksheet into list of KOL records."""
        records = []
        
        # Read header row (first row)
        headers = []
        for cell in sheet[1]:
            header = str(cell.value).strip().lower().replace(' ', '_') if cell.value else ''
            headers.append(header)
        
        # Map headers to our field names
        field_mapping = cls._create_field_mapping(headers)
        
        # Read data rows (skip header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            record = cls._parse_row(row, headers, field_mapping, row_idx)
            if record:  # Only add if we got valid data
                records.append(record)
        
        return records
    
    @classmethod
    def _create_field_mapping(cls, headers: List[str]) -> Dict[int, str]:
        """Create mapping from column index to field name."""
        mapping = {}
        
        for idx, header in enumerate(headers):
            if header in cls.COLUMN_MAPPINGS:
                mapping[idx] = cls.COLUMN_MAPPINGS[header]
        
        return mapping
    
    @classmethod
    def _parse_row(
        cls, 
        row: tuple, 
        headers: List[str], 
        field_mapping: Dict[int, str],
        row_number: int
    ) -> Optional[Dict[str, Any]]:
        """Parse a single row into a KOL record."""
        record = {}
        
        for idx, value in enumerate(row):
            if idx not in field_mapping:
                continue
            
            field_name = field_mapping[idx]
            parsed_value = cls._parse_value(value, field_name)
            
            if parsed_value is not None:
                record[field_name] = parsed_value
        
        # Generate ID if missing
        if 'id' not in record:
            record['id'] = str(row_number - 1)
        
        # Only return if we have minimum required fields
        if 'name' in record:
            return record
        
        return None
    
    @classmethod
    def _parse_value(cls, value: Any, field_name: str) -> Any:
        """Parse and convert a cell value to appropriate type."""
        if value is None:
            return None
        
        # Numeric fields
        if field_name in ['publicationsCount', 'hIndex', 'citations']:
            try:
                return int(float(value)) if value else None
            except (ValueError, TypeError):
                return None
        
        # String fields
        return str(value).strip() if value else None



