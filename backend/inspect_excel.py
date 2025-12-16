"""Quick script to inspect Excel file structure."""
import openpyxl
from pathlib import Path

excel_file = Path("../Vitiligo_kol_csv_29_07_2024_drug_and_kol_standardized.xlsx")

if excel_file.exists():
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    print()
    
    sheet = wb.active
    print(f"Active sheet: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")
    print()
    
    print("First 10 rows:")
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        print(f"{i}: {row[:10]}")  # Show first 10 columns



